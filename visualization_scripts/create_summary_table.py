#!/usr/bin/env python3
"""
Create a summary table with weighted averages for columns 1-18, grouped by video_id.
Uses frame_activation_norm as weights and excludes "NA" values.
"""

import pandas as pd
import numpy as np
import argparse
import sys
from openpyxl import load_workbook

def create_summary_table(excel_path):
    """
    Create a summary table with weighted averages and add it as a new sheet.
    
    Args:
        excel_path: Path to the Excel workbook
    """
    # Read the Excel file
    # Use keep_default_na=False to prevent pandas from converting "NA" strings to NaN
    print(f"Reading Excel file: {excel_path}")
    df = pd.read_excel(excel_path, sheet_name=0, keep_default_na=False, na_values=[''])
    
    print(f"Data shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    
    # Validate that columns 1-18 and background only contain 0, 1, 2, 3, or "NA"
    # Also check that empty values (NaN) only occur when ALL rows for a video_id are empty
    print("\nValidating data...")
    valid_values = {0, 1, 2, 3, "NA"}
    columns_to_check = [str(i) for i in range(1, 19)] + ['background']
    
    # Get unique video_ids for validation and later use
    unique_video_ids = sorted(df['video_id'].unique())
    
    for col in columns_to_check:
        if col not in df.columns:
            continue
        
        # Check each video_id separately
        for video_id in unique_video_ids:
            video_rows = df[df['video_id'] == video_id]
            
            # Check for invalid values (not in valid set)
            for idx, value in video_rows[col].items():
                # Skip NaN/empty values for now - we'll check them separately
                if pd.isna(value) or value == '':
                    continue
                
                # Convert numeric strings to numbers, and handle numeric values
                is_valid = False
                if value == "NA":
                    is_valid = True
                else:
                    # Try to convert to numeric and check if it's 0, 1, 2, or 3
                    try:
                        num_value = float(value)
                        if num_value.is_integer() and int(num_value) in [0, 1, 2, 3]:
                            is_valid = True
                    except (ValueError, TypeError):
                        pass
                
                if not is_valid:
                    print(f"ERROR: Invalid value found in column '{col}' at row {idx + 2} (Excel row {idx + 2}): {value} (type: {type(value).__name__})")
                    print(f"Valid values are: 0, 1, 2, 3, or 'NA'")
                    print("Exiting without creating summary table.")
                    sys.exit(1)
            
            # Check for mixed empty/filled values within the same video_id
            # Empty means NaN or empty string
            has_empty = (video_rows[col].isna() | (video_rows[col] == '')).any()
            has_filled = (~video_rows[col].isna() & (video_rows[col] != '')).any()
            
            if has_empty and has_filled:
                # Find the first empty row for error message
                empty_rows = video_rows[video_rows[col].isna() | (video_rows[col] == '')]
                first_empty_idx = empty_rows.index[0]
                print(f"ERROR: Mixed empty and filled values found in column '{col}' for video_id {video_id}")
                print(f"Empty value found at row {first_empty_idx + 2} (Excel row {first_empty_idx + 2})")
                print("All rows for a video_id must either all be filled or all be empty.")
                print("Exiting without creating summary table.")
                sys.exit(1)
    
    print("Validation passed: All values in columns 1-18 and background are valid.")
    print(f"Found {len(unique_video_ids)} unique video_ids: {unique_video_ids}")
    
    # Create summary data structure with video_id, pred_category, and true_category
    summary_data = {'video_id': unique_video_ids}
    
    # Add pred_category and true_category (should be same for all rows of same video_id)
    pred_categories = []
    true_categories = []
    for video_id in unique_video_ids:
        video_data = df[df['video_id'] == video_id]
        # Get the first value (should be same for all rows)
        pred_categories.append(video_data['pred_category'].iloc[0])
        true_categories.append(video_data['true_category'].iloc[0])
    
    summary_data['pred_category'] = pred_categories
    summary_data['true_category'] = true_categories
    
    # Columns to calculate weighted averages for (1-18)
    num_columns = [str(i) for i in range(1, 19)]
    
    # Calculate weighted average for each column, grouped by video_id
    # Each video_id gets its own weighted average calculated from all rows for that video_id
    for col in num_columns:
        if col in df.columns:
            print(f"Processing column {col}...")
            weighted_avgs = []
            for video_id in unique_video_ids:
                # Filter data for this video_id
                video_data = df[df['video_id'] == video_id].copy()
                
                # Filter out rows where column value is "NA" or not numeric
                valid_mask = (video_data[col] != 'NA') & (pd.to_numeric(video_data[col], errors='coerce').notna())
                valid_data = video_data[valid_mask]
                
                if len(valid_data) > 0:
                    # Get values and weights
                    values = pd.to_numeric(valid_data[col], errors='coerce')
                    weights = valid_data['frame_activation_norm']
                    
                    # Calculate weighted average: sum(value * weight) / sum(weight)
                    weighted_avg = np.average(values, weights=weights)
                    weighted_avgs.append(weighted_avg)
                else:
                    # No valid data for this video_id and column
                    weighted_avgs.append(np.nan)
            
            summary_data[col] = weighted_avgs
        else:
            print(f"Warning: Column {col} not found in data")
    
    # Create summary DataFrame
    summary_df = pd.DataFrame(summary_data)
    
    # Load the workbook to add a new sheet
    print(f"\nAdding video_summary sheet to workbook...")
    book = load_workbook(excel_path)
    
    # Remove the video_summary sheet if it already exists
    if 'video_summary' in book.sheetnames:
        print("Removing existing 'video_summary' sheet...")
        del book['video_summary']
    
    # Create a new sheet
    book.create_sheet('video_summary')
    
    # Save the workbook structure
    book.save(excel_path)
    
    # Write the summary DataFrame to the new sheet
    # Ensure column order: video_id, pred_category, true_category, then columns 1-18
    column_order = ['video_id', 'pred_category', 'true_category'] + num_columns
    summary_df_ordered = summary_df[[col for col in column_order if col in summary_df.columns]]
    
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        summary_df_ordered.to_excel(writer, sheet_name='video_summary', index=False)
    
    print(f"\nVideo summary table created successfully!")
    print(f"Video summary table shape: {summary_df_ordered.shape}")
    print(f"\nFirst few rows:")
    print(summary_df_ordered.head())
    
    # Create species_summary: average of numbered columns grouped by true_category
    print(f"\nCreating species_summary sheet...")
    unique_true_categories = sorted(df['true_category'].unique())
    
    species_summary_data = {'true_category': unique_true_categories}
    
    # Calculate average for each numbered column, grouped by true_category
    for col in num_columns:
        if col in df.columns:
            print(f"Processing column {col} for species summary...")
            category_avgs = []
            for true_cat in unique_true_categories:
                # Filter data for this true_category
                category_data = df[df['true_category'] == true_cat].copy()
                
                # Filter out rows where column value is "NA" or not numeric
                valid_mask = (category_data[col] != 'NA') & (pd.to_numeric(category_data[col], errors='coerce').notna())
                valid_data = category_data[valid_mask]
                
                if len(valid_data) > 0:
                    # Get numeric values and calculate mean (ignoring NA)
                    values = pd.to_numeric(valid_data[col], errors='coerce')
                    category_avg = values.mean()
                    category_avgs.append(category_avg)
                else:
                    # No valid data for this category and column
                    category_avgs.append(np.nan)
            
            species_summary_data[col] = category_avgs
    
    # Create species summary DataFrame
    species_summary_df = pd.DataFrame(species_summary_data)
    
    # Reload workbook to add species_summary sheet
    book = load_workbook(excel_path)
    
    # Remove species_summary sheet if it already exists
    if 'species_summary' in book.sheetnames:
        print("Removing existing 'species_summary' sheet...")
        del book['species_summary']
    
    # Create a new sheet
    book.create_sheet('species_summary')
    
    # Save the workbook structure
    book.save(excel_path)
    
    # Write the species summary DataFrame to the new sheet
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        species_summary_df.to_excel(writer, sheet_name='species_summary', index=False)
    
    print(f"\nSpecies summary table created successfully!")
    print(f"Species summary table shape: {species_summary_df.shape}")
    print(f"\nSpecies summary:")
    print(species_summary_df)
    
    return summary_df_ordered

def main():
    parser = argparse.ArgumentParser(
        description='Create a summary table with weighted averages for columns 1-18, grouped by video_id'
    )
    parser.add_argument(
        'excel_path',
        type=str,
        help='Path to the Excel workbook'
    )
    
    args = parser.parse_args()
    
    try:
        create_summary_table(args.excel_path)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()

